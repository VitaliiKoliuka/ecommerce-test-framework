import json
import csv
from openpyxl import load_workbook
from datetime import datetime


class DataProvider:

    @staticmethod
    def read_excel_data(filepath, sheetname):
        wb = load_workbook(filepath)
        sheet = wb[sheetname]
        data = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            data.append(row)

        return data

    @staticmethod
    def read_json_data(filepath):
        with open(filepath, 'r') as f:
            data_list = json.load(f)

        return [(item,) for item in data_list]

    @staticmethod
    def read_csv_data(filepath):
        data = []

        with open(filepath, 'r') as f:
            reader = csv.reader(f)
            next(reader)

            for row in reader:
                data.append(tuple(row))

        return data

def validate_cart_dates_within_range(cart_dates,start,end):
    fmt = "%Y-%m-%d"
    start = datetime.strptime(start, fmt)
    end = datetime.strptime(end, fmt)
    for d in cart_dates:
        dt = datetime.strptime(d[:10], fmt)
        if not (start <= dt <= end):
            return False
    return True